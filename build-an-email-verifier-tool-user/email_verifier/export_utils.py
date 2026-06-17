from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from fpdf import FPDF
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


GREEN_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
RED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
GRAY_FILL = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFD8A8", end_color="FFD8A8", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_COL = "Final Status"


def _status_fill(val: str) -> PatternFill:
    if val in ("OK", "Valid"):
        return GREEN_FILL
    elif val in ("Catch-All", "Risky", "Unknown"):
        return YELLOW_FILL
    elif val in ("Invalid", "Disposable"):
        return RED_FILL
    elif val == "Duplicate":
        return GRAY_FILL
    return PatternFill()


def _apply_column_widths(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, 1):
        max_len = max(
            df[col].astype(str).map(len).max() if len(df) else 0,
            len(str(col)),
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 40)


def _style_header(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _color_rows(ws, df: pd.DataFrame) -> None:
    if STATUS_COL not in df.columns:
        return
    col_idx = df.columns.get_loc(STATUS_COL) + 1
    for row_idx in range(2, len(df) + 2):
        status_val = str(ws.cell(row=row_idx, column=col_idx).value)
        fill = _status_fill(status_val)
        for c_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=c_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


def dataframe_to_xlsx_preserve(
    all_sheets: dict[str, pd.DataFrame],
    verification_df: pd.DataFrame,
) -> bytes:
    base_name = "Email Verification Report"
    sheet_name = base_name
    counter = 2
    while sheet_name in all_sheets:
        sheet_name = f"{base_name} ({counter})"
        counter += 1

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for orig_name, orig_df in all_sheets.items():
            orig_df.to_excel(writer, index=False, sheet_name=orig_name)

        verification_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _style_header(ws, verification_df)
        _apply_column_widths(ws, verification_df)
        _color_rows(ws, verification_df)
        _build_summary_sheet(writer, verification_df)
        _build_separate_sheets(writer, verification_df)

    return output.getvalue()


def dataframe_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Verification Results")
        ws = writer.sheets["Verification Results"]
        _style_header(ws, df)
        _apply_column_widths(ws, df)
        _color_rows(ws, df)

        _build_summary_sheet(writer, df)
        _build_separate_sheets(writer, df)

    return output.getvalue()


def _build_summary_sheet(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    total = len(df)
    ok_count = int((df[STATUS_COL] == "OK").sum())
    catchall_count = int((df[STATUS_COL] == "Catch-All").sum())
    risky_count = int((df[STATUS_COL] == "Risky").sum())
    invalid_count = int((df[STATUS_COL] == "Invalid").sum())
    disposable_count = int((df[STATUS_COL] == "Disposable").sum())
    unknown_count = int((df[STATUS_COL] == "Unknown").sum())
    duplicate_count = int((df[STATUS_COL] == "Duplicate").sum())

    summary_data = {
        "Metric": [
            "Total Emails",
            "OK (Send)",
            "Catch-All (Send carefully)",
            "Risky (Review first)",
            "Invalid (Do not send)",
            "Disposable (Do not send)",
            "Unknown (Usually avoid)",
            "Duplicate (Remove)",
            "",
            "Generated",
        ],
        "Value": [
            total,
            ok_count,
            catchall_count,
            risky_count,
            invalid_count,
            disposable_count,
            unknown_count,
            duplicate_count,
            "",
            datetime.now().strftime("%d-%b-%Y %I:%M %p"),
        ],
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, index=False, sheet_name="Summary")

    ws = writer.sheets["Summary"]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    for i, col in enumerate(summary_df.columns, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    fill_map = {
        2: GREEN_FILL,
        3: YELLOW_FILL,
        4: YELLOW_FILL,
        5: RED_FILL,
        6: RED_FILL,
        7: YELLOW_FILL,
        8: GRAY_FILL,
    }
    for row_num, fill in fill_map.items():
        for c in range(1, 3):
            ws.cell(row=row_num + 1, column=c).fill = fill
            ws.cell(row=row_num + 1, column=c).border = THIN_BORDER

    ws.cell(row=12, column=1).value = "Color Coding:"
    ws.cell(row=12, column=1).font = Font(bold=True)
    ws.cell(row=13, column=1).value = "🟢 Green  = OK / Valid"
    ws.cell(row=13, column=1).fill = GREEN_FILL
    ws.cell(row=14, column=1).value = "🟡 Yellow = Catch-All / Risky / Unknown"
    ws.cell(row=14, column=1).fill = YELLOW_FILL
    ws.cell(row=15, column=1).value = "🔴 Red    = Invalid / Disposable"
    ws.cell(row=15, column=1).fill = RED_FILL
    ws.cell(row=16, column=1).value = "⬜ Gray   = Duplicate"
    ws.cell(row=16, column=1).fill = GRAY_FILL


def _build_separate_sheets(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    mapping = {
        "OK_Emails": "OK",
        "CatchAll_Risky_Emails": ("Catch-All", "Risky"),
        "Invalid_Emails": ("Invalid", "Disposable"),
        "Unknown_Emails": "Unknown",
        "Duplicate_Emails": "Duplicate",
    }
    for sheet_name, status_filter in mapping.items():
        if isinstance(status_filter, tuple):
            subset = df[df[STATUS_COL].isin(status_filter)]
        else:
            subset = df[df[STATUS_COL] == status_filter]
        if subset.empty:
            continue
        subset.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        _style_header(ws, subset)
        _apply_column_widths(ws, subset)
        _color_rows(ws, subset)


def dataframe_valid_xlsx_bytes(df: pd.DataFrame) -> bytes:
    return _filtered_xlsx(df, ("OK",))


def dataframe_risky_xlsx_bytes(df: pd.DataFrame) -> bytes:
    return _filtered_xlsx(df, ("Catch-All", "Risky"))


def dataframe_invalid_xlsx_bytes(df: pd.DataFrame) -> bytes:
    return _filtered_xlsx(df, ("Invalid", "Disposable"))


def _filtered_xlsx(df: pd.DataFrame, statuses: tuple[str, ...]) -> bytes:
    output = io.BytesIO()
    subset = df[df[STATUS_COL].isin(statuses)].copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        subset.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]
        _style_header(ws, subset)
        _apply_column_widths(ws, subset)
        _color_rows(ws, subset)
    return output.getvalue()


def dataframe_to_pdf_bytes(df: pd.DataFrame) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Email Verification Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", align="C", new_x="LMARGIN", new_y="NEXT")

    total = len(df)
    ok_c = int((df[STATUS_COL] == "OK").sum())
    risky_c = int(df[STATUS_COL].isin(("Catch-All", "Risky")).sum())
    bad_c = int(df[STATUS_COL].isin(("Invalid", "Disposable")).sum())
    unknown_c = int((df[STATUS_COL] == "Unknown").sum())

    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, f"Total: {total}  |  OK: {ok_c}  |  Risky: {risky_c}  |  Bad: {bad_c}  |  Unknown: {unknown_c}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    cols = list(df.columns)
    display_cols = [c for c in cols if c not in ("LinkedIn URL", "SMTP Response")]
    if not display_cols:
        display_cols = cols[:8]

    col_width = max(22, int(180 / len(display_cols)))

    pdf.set_font("Helvetica", "B", 6)
    pdf.set_fill_color(68, 114, 196)
    pdf.set_text_color(255, 255, 255)
    for col in display_cols:
        pdf.cell(col_width, 7, _sanitize(col)[:18], border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    pdf.set_font("Helvetica", "", 5)
    for _, row in df.iterrows():
        fs = str(row.get(STATUS_COL, ""))
        if fs == "OK":
            pdf.set_fill_color(212, 237, 222)
        elif fs in ("Catch-All", "Risky"):
            pdf.set_fill_color(255, 243, 205)
        elif fs in ("Invalid", "Disposable"):
            pdf.set_fill_color(248, 215, 218)
        elif fs == "Unknown":
            pdf.set_fill_color(255, 216, 168)
        else:
            pdf.set_fill_color(226, 227, 229)

        if pdf.get_y() > 265:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 6)
            pdf.set_fill_color(68, 114, 196)
            pdf.set_text_color(255, 255, 255)
            for col in display_cols:
                pdf.cell(col_width, 7, col[:18], border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 5)

        for col in display_cols:
            val = _sanitize(str(row.get(col, "")))[:22]
            pdf.cell(col_width, 5, val, border=1, align="C", fill=True)
        pdf.ln()

    return bytes(pdf.output(dest="S"))


def _sanitize(text: str) -> str:
    return text.encode("ascii", errors="replace").decode("ascii")
