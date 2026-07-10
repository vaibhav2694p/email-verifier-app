from io import BytesIO
from typing import List

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExportService:
    @staticmethod
    def to_csv(df: pd.DataFrame) -> bytes:
        return df.to_csv(index=False).encode("utf-8")

    @staticmethod
    def to_excel(df: pd.DataFrame) -> bytes:
        buf = BytesIO()

        status_order = [
            "Valid", "Likely Valid", "Invalid", "Catch-All", "Disposable",
            "Role Account", "Abuse", "Unknown", "Temporary Failure", "Do Not Mail",
            "SMTP Blocked", "Greylisted", "Duplicate", "Risky",
        ]

        sheets = {
            "All Results": df.copy(),
        }
        if "verification_status" in df.columns:
            for status in status_order:
                subset = df[df["verification_status"] == status].copy()
                if not subset.empty:
                    sheets[_sheet_name(status)] = subset

            if "do_not_mail" in df.columns:
                subset = df[df["do_not_mail"] == True].copy()
                if not subset.empty:
                    sheets["Do Not Mail"] = subset

            if "is_duplicate" in df.columns:
                subset = df[df["is_duplicate"] == True].copy()
                if not subset.empty:
                    sheets["Duplicate List"] = subset

        domain_summary = ExportService.to_domain_summary(df)
        if not domain_summary.empty:
            sheets["Domain Summary"] = domain_summary

        provider_summary = ExportService.to_provider_summary(df)
        if not provider_summary.empty:
            sheets["Provider Summary"] = provider_summary

        run_summary = pd.DataFrame([{
            "Metric": "Total Emails",
            "Value": len(df),
        }])
        if "verification_status" in df.columns:
            for status in status_order:
                count = int((df["verification_status"] == status).sum())
                run_summary = pd.concat([run_summary, pd.DataFrame([{"Metric": status, "Value": count}])], ignore_index=True)
        if "is_duplicate" in df.columns:
            dup_count = int(df["is_duplicate"].sum())
            run_summary = pd.concat([run_summary, pd.DataFrame([{"Metric": "Duplicates", "Value": dup_count}])], ignore_index=True)
        sheets["Run Summary"] = run_summary

        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheets.items():
                sheet_df = sheet_df.copy()
                for col in sheet_df.columns:
                    if sheet_df[col].dtype == "object":
                        sheet_df[col] = sheet_df[col].astype(str)
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
                ws = writer.sheets[sheet_name[:31]]
                ExportService._format_worksheet(ws, list(sheet_df.columns))
                ExportService._apply_status_coloring(ws, sheet_df)

        return buf.getvalue()

    @staticmethod
    def to_filtered_csv(df: pd.DataFrame, status: str) -> bytes:
        if status and "verification_status" in df.columns:
            filtered = df[df["verification_status"] == status].copy()
        else:
            filtered = df.copy()
        return filtered.to_csv(index=False).encode("utf-8")

    @staticmethod
    def to_provider_summary(df: pd.DataFrame) -> pd.DataFrame:
        provider_col = "mail_hosting_provider" if "mail_hosting_provider" in df.columns else "mx_provider"
        if provider_col not in df.columns:
            return pd.DataFrame()
        summary = df.groupby(provider_col).agg(total=(provider_col, "count")).reset_index()
        if "verification_status" in df.columns:
            status_counts = df.groupby([provider_col, "verification_status"]).size().unstack(fill_value=0)
            for col in status_counts.columns:
                summary[col] = summary[provider_col].map(status_counts[col]).fillna(0).astype(int)
        if "verification_score" in df.columns:
            scores = df.groupby(provider_col)["verification_score"].mean().round(1).reset_index()
            scores.columns = [provider_col, "avg_score"]
            summary = summary.merge(scores, on=provider_col, how="left")
        return summary.sort_values("total", ascending=False).reset_index(drop=True)

    @staticmethod
    def to_domain_summary(df: pd.DataFrame) -> pd.DataFrame:
        if "domain" not in df.columns:
            return pd.DataFrame()

        summary = df.groupby("domain").agg(
            total=("domain", "count"),
        ).reset_index()

        if "verification_status" in df.columns:
            status_counts = df.groupby(["domain", "verification_status"]).size().unstack(fill_value=0)
            for col in status_counts.columns:
                summary[col] = summary["domain"].map(status_counts[col]).fillna(0).astype(int)

        if "verification_score" in df.columns:
            score_agg = df.groupby("domain")["verification_score"].agg(["mean", "min", "max"]).reset_index()
            score_agg.columns = ["domain", "avg_score", "min_score", "max_score"]
            summary = summary.merge(score_agg, on="domain", how="left")
            summary["avg_score"] = summary["avg_score"].round(1)

        if "mx_provider" in df.columns:
            provider_agg = df.groupby("domain")["mx_provider"].first().reset_index()
            summary = summary.merge(provider_agg, on="domain", how="left")

        summary = summary.sort_values("total", ascending=False).reset_index(drop=True)
        return summary

    @staticmethod
    def _format_worksheet(ws, headers: List[str]):
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin", color="D0D5DD"),
            right=Side(style="thin", color="D0D5DD"),
            top=Side(style="thin", color="D0D5DD"),
            bottom=Side(style="thin", color="D0D5DD"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.freeze_panes = "A2"

        if headers:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

        ExportService._auto_column_widths(ws)

    @staticmethod
    def _auto_column_widths(ws):
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    val = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(val))
                except Exception:
                    pass
            adjusted_width = min(max_length + 3, 50)
            adjusted_width = max(adjusted_width, 10)
            ws.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def _apply_status_coloring(ws, df):
        status_colors = {
            "Valid": "DCFCE7",
            "Likely Valid": "DBEAFE",
            "Risky": "FEF3C7",
            "Invalid": "FEE2E2",
            "Unknown": "F1F5F9",
            "Catch-All": "FEF3C7",
            "Disposable": "FEE2E2",
            "Role Account": "FEF3C7",
            "Abuse": "FEE2E2",
            "Temporary Failure": "E0F2FE",
            "Do Not Mail": "FECACA",
            "SMTP Blocked": "E5E7EB",
            "Greylisted": "E0F2FE",
            "Duplicate": "E5E7EB",
        }

        if "verification_status" not in df.columns:
            return

        status_col_idx = None
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == "verification_status":
                status_col_idx = col_idx
                break

        if status_col_idx is None:
            return

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=status_col_idx)
            status_val = str(cell.value) if cell.value is not None else ""
            color = status_colors.get(status_val)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill


def _sheet_name(status: str) -> str:
    mapping = {
        "Catch-All": "Catch-All",
        "Role Account": "Role Accounts",
        "Temporary Failure": "Temporary Failures",
        "Do Not Mail": "Do Not Mail",
    }
    return mapping.get(status, status)
